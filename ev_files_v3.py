# -*- coding: utf-8 -*-
"""
Created on Thu Jun  9 10:01:02 2016

@author: bschloss
"""
import argparse as ap
import codecs
import os
import openpyxl as xl
import numpy as np
import string


def charsonly(s):
	acceptable = string.ascii_letters + "1234567890'- "
    	new = ''
    	for char in s:  
		if char in acceptable:
			new += char
	return new 
 
def l2t(inp):
	txt = ''
	for line in inp:
		txt += line + '\n'
	return txt
 
def normalize(lines):
    vals = []    
    new = []
    for line in lines:
        vals.append(float(line.split('\t')[-1].rstrip('\n')))
    mean = np.mean(vals)
    std = np.std(vals)
    for i in range(len(lines)):
        val = "{0:.2f}".format((vals[i]-mean)/std)
        new.append('\t'.join([lines[i].split('\t')[0],lines[i].split('\t')[1],val]))
    return new
        
def quadratic(lines):  
    new = []
    for i in range(len(lines)):
        line = lines[i]
        val = "{0:.2f}".format(float(line.split('\t')[-1].rstrip('\n'))**2)
        new.append('\t'.join([line.split('\t')[0],line.split('\t')[1],val]))
    return normalize(new)
    
def main():
    
    parser = ap.ArgumentParser(description='Create dictionary')
    parser.add_argument('pardir', metavar='in', type=str, 
                       help='Participant directory')
    parser.add_argument('text_data', metavar='td', type=str, 
                       help='compiled_Text_data')
    parser.add_argument('randperm', metavar='rp', type =str,
                        help='Random orders of texts for 30 pars')
    args = parser.parse_args()
    
    pardir = args.pardir  
    text_data_path = args.text_data     
    randperm_path =args.randperm
         
    EPRIME_DIR = pardir.rstrip('/') + '/E-Prime/'
    EV_DIR = pardir.rstrip('/') + '/EV3/'
    EDF_DIR = pardir.rstrip('/') + '/EDF/'
    
    num2txt = {'1':'/home/bschloss/pul8_read/Final_Texts/Txt/Mars_Text_310.txt',
           '2':'/home/bschloss/pul8_read/Final_Texts/Txt/Supertanker_Text_302.txt',
           '3':'/home/bschloss/pul8_read/Final_Texts/Txt/Math_Text_306.txt',
           '4':'/home/bschloss/pul8_read/Final_Texts/Txt/GPS_Text_307.txt',
           '5':'/home/bschloss/pul8_read/Final_Texts/Txt/Circuit_Text_302.txt'}
    td = xl.load_workbook(text_data_path) 
    rp = [line.rstrip('\r\n') for line in open(randperm_path,'r').readlines()]
    texts = [(rp[i],num2txt[str(rp[i])]) for i in range(5*(int(pardir.rstrip('/')[-2:])-1),5*int(pardir.rstrip('/')[-2:]))]
    runs = ['Run' + str(i) for i in range(1,6)]
    
    eprime_dirs = [EPRIME_DIR + run + '/' for run in runs]
    edf_dirs = [EDF_DIR + run + '/' for run in runs]
    
    for d in eprime_dirs:
        if '2006/E-Prime/Run5' not in d:
            f = d + os.listdir(d)[0]
            runnum = int(d[-2])
            f = codecs.open(f,'r',encoding='utf-16-le').readlines()
            
            instructions = ''
            questions = ''
            
            for line in f:
                
                #Waiting for Trigger
                #Time Zero 
                if u"WaitForTrigger.OffsetTime:" in line:
                    print line
                    timezero = float(line.lstrip("WaitForTrigger.OffsetTime: ").rstrip('\r\n'))
            
            for line in f:
            
                #Duration of 2 seconds
                #Instruction Screen
                if u"NextText.OnsetTime:" in line:
                    print line
                    onset = float(line.lstrip("\tNextText.OnsetTime: ").rstrip('\r\n')) - timezero
                    onset = onset/1000
                    duration = 2.0
                    instructions += str(onset) + '\t' + str(duration) + '\t' + str(1.0) +'\n'
            
                #After all of the sentences and 6 seconds of fixation
                #Instruction Screen
                if u"TextEnd.OnsetTime:" in line:
                    print line
                    onset = float(line.lstrip("\tTextEnd.OnsetTime: ").rstrip('\r\n')) - timezero
                    onset = onset/1000
                    duration = 5.0
                    instructions += str(onset) + '\t' + str(duration) + '\t' + str(1.0) + '\n'
                
            
                #Model onset of first question and RTTime (Offset) of last question as a single condition
                
                if u"Questions.OnsetTime:" in line:
                    q_onset = float(line.lstrip("\t\tQuestions.OnsetTime:: ").rstrip('\r\n')) - timezero
                if u"Questions.RTTime:" in line:
                    q_offset =float(line.lstrip('\t\tQuestions.RTTime: ').rstrip('\r\n'))
                    q_duration = q_offset - timezero - q_onset
                    q_onset = q_onset/1000.0
                    q_duration = q_duration/1000.0
                    questions += "{0:.2f}".format(q_onset) + '\t' + "{0:.2f}".format(q_duration) + '\t' + "{0:.2f}".format(float(1)) + '\n'
                
            
                #Final instructions, Duration 3 seconds
                if u"Goodbye.OnsetTime:" in line:
                    print line
                    onset = float(line.lstrip("Goodbye.OnsetTime: ").rstrip('\r\n')) - timezero
                    onset = onset/1000.0
                    duration = 3.0
                    instructions += str(onset) + '\t' + str(duration) + '\t' + str(1.0) 
                
                if u"TextEnd.OnsetTime:" in line:
                    text_end = float(line.lstrip("\tTextEnd.OnsetTime: ").rstrip('\r\n')) - timezero
                    text_end_vol_num = int(text_end/400.00)
                    open(''.join([EV_DIR,'Run',str(runnum),'/','text_end_time.run00',str(runnum),'.txt']),'w').write("{0:.2f}".format(text_end/1000.0))
                    open(''.join([EV_DIR,'Run',str(runnum),'/','text_end_vol_num.run00',str(runnum),'.txt']),'w').write(str(text_end_vol_num))
        
            if instructions:
		open(''.join([EV_DIR,'Run',str(runnum),'/','instructions.run00',str(runnum),'.txt']),'w').write(instructions)
                first_instructions = instructions.split('\n')[0]
                open(''.join([EV_DIR,'Run',str(runnum),'/','first_instructions.run00',str(runnum),'.txt']),'w').write(first_instructions)
            if questions:
		open(''.join([EV_DIR,'Run',str(runnum),'/','questions.run00',str(runnum),'.txt']),'w').write(questions)
       
    for d in edf_dirs:
        if ('021/EDF/Run4/' not in d or '2021/EDF/Run4/' in d)  and '2005/EDF/Run2' not in d and '2006/EDF/Run5' not in d and '2014/EDF/Run5' not in d:
            runnum = int(d[-2])
            
            textnum,text = texts[runnum-1]
            text_data = td.get_sheet_by_name(text.split('/')[-1].replace('.txt',''))
            
            evs = {}
            binaries = ['Content_Word',
                        'Target_Word',
                        'Content_or_Target_Word', 
                        'No_Interest',
                        'Content_Word_with_Regressions', 
                        'Target_Word_with_Regressions',
                        'Content_or_Target_Word_with_Regressions',
                        'No_Interest_without_Regressions',
                        'Regressions',
                        'Content_Word_Regressions',
                        'Target_Word_Regressions',
                        'Content_or_Target_Word_Regressions',
                        'No_Interest_Regressions']
                    
            for key in binaries:
                evs[key] = ''
                evs['_'.join([key,'with_Words'])] = ''
                
            evs['Position_Target'] = {'text':'','column':4}
            evs['Position_Content'] = {'text':'','column':4}
            evs['Position_Content_or_Target'] = {'text':'','column':4}
            evs['Position_Target_with_Regressions'] = {'text':'','column':4}
            evs['Position_Content_with_Regressions'] = {'text':'','column':4}
            evs['Position_Content_or_Target_with_Regressions'] = {'text':'','column':4}
            
            col = 10
            while(text_data.cell(row=1,column=col).value!=None):
                evs['_'.join([text_data.cell(row=1,column=col).value,'Target'])] = {'column':col,'text':''}
                evs['_'.join([text_data.cell(row=1,column=col).value,'Content'])] = {'column':col,'text':''}
                evs['_'.join([text_data.cell(row=1,column=col).value,'Content_or_Target'])] = {'column':col,'text':''}
                evs['_'.join([text_data.cell(row=1,column=col).value,'Target_with_Regressions'])] = {'column':col,'text':''}
                evs['_'.join([text_data.cell(row=1,column=col).value,'Content_with_Regressions'])] = {'column':col,'text':''}
                evs['_'.join([text_data.cell(row=1,column=col).value,'Content_or_Target_with_Regressions'])] = {'column':col,'text':''}
                col+=1        
    
            edf = d + os.listdir(d)[0]
            print edf,text
            edfws = xl.load_workbook(edf).active
            
            fix_onset_col = 1
            while(str(edfws.cell(row=1,column = fix_onset_col).value)!='CURRENT_FIX_START'):
                fix_onset_col += 1
            fix_dur_col = 1
            while(str(edfws.cell(row=1,column = fix_dur_col).value)!='CURRENT_FIX_DURATION'):
                fix_dur_col += 1
            area_id_col = 1
            while((str(edfws.cell(row=1,column = area_id_col).value)!='CURRENT_FIX_INTEREST_AREA_ID' and str(edfws.cell(row=1,column= area_id_col).value)!='CURRENT_FIX_INTEREST_AREA_INDEX')):
                area_id_col += 1
            sent_onset_col = 1
            while(str(edfws.cell(row=1,column = sent_onset_col).value)!='TRUEOnset'):
                sent_onset_col += 1
            sent_offset_col = 1
            while(str(edfws.cell(row=1,column = sent_offset_col).value)!='TRUEOffset'):
                sent_offset_col += 1
            sent_dur_col = 1
            while(str(edfws.cell(row=1,column = sent_dur_col).value)!='SentenceRT'):
                sent_dur_col += 1
            sent_id_col = 1
            while(str(edfws.cell(row=1,column = sent_id_col).value)!='SentenceID'):
                sent_id_col += 1
            rsl_col = 1 #Recording Session Label
            while(str(edfws.cell(row=1,column = rsl_col).value) not in ['RECORDING SESSION LABEL','RECORDING_SESSION_LABEL']):
                rsl_col += 1
            edf_row = 1
            while(str(edfws.cell(row=edf_row,column = sent_id_col).value)[:2]!='t.'):
                edf_row += 1
            
            if not os.path.isfile(''.join([EV_DIR,'Run',str(runnum),'/','first_instructions.run00',str(runnum),'.txt'])):
                fi_onset = float(edfws.cell(row = edf_row, column = sent_onset_col).value)/1000 - 8.5
                first_instructions = "{0:.2f}".format(fi_onset) + '\t2.0\t1.0'
                open(''.join([EV_DIR,'Run',str(runnum),'/','first_instructions.run00',str(runnum),'.txt']),'w').write(first_instructions)
                
            if not os.path.isfile(''.join([EV_DIR,'Run',str(runnum),'/','text_end_time.run00',str(runnum),'.txt'])):
                sent_nums = []
                edf_row2 = edf_row
                while(edfws.cell(row=edf_row2,column=rsl_col).value!=None):
                    if 't' in str(edfws.cell(row=edf_row2,column=sent_id_col).value):
                        sent_nums.append(int(str(edfws.cell(row=edf_row2,column=sent_id_col).value).split('.')[-1]))
                    edf_row2 += 1   
                max_sent_num = max(sent_nums)
                edf_row2 = edf_row
                while(str(edfws.cell(row=edf_row2,column=sent_id_col).value).split('.')[-1] != str(max_sent_num)):
                    edf_row2 += 1
                text_end_time = float(edfws.cell(row=edf_row2,column=sent_offset_col).value)/1000 + 6
                open(''.join([EV_DIR,'Run',str(runnum),'/','text_end_time.run00',str(runnum),'.txt']),'w').write("{0:.2f}".format(text_end_time))
                
            if not os.path.isfile(''.join([EV_DIR,'Run',str(runnum),'/','text_end_vol_num.run00',str(runnum),'.txt'])):
                text_end_vol_num = int(text_end_time/.400)
                open(''.join([EV_DIR,'Run',str(runnum),'/','text_end_vol_num.run00',str(runnum),'.txt']),'w').write(str(text_end_vol_num))
                
            if not os.path.isfile(''.join([EV_DIR,'Run',str(runnum),'/','questions.run00',str(runnum),'.txt'])):
                edf_row2 = edf_row
                question_onsets = []
                question_offsets = []
                while(edfws.cell(row=edf_row2,column=rsl_col).value!=None):
                    if 't' in str(edfws.cell(row=edf_row2,column=sent_id_col).value):
                        sent_nums.append(int(str(edfws.cell(row=edf_row2,column=sent_id_col).value).split('.')[-1]))
                    edf_row2 += 1
                max_sent_num = max(sent_nums)
                edf_row2 = edf_row
                while(str(edfws.cell(row=edf_row2,column=sent_id_col).value).split('.')[-1] != str(max_sent_num)):
                    edf_row2 += 1
                while(str(edfws.cell(row=edf_row2,column=sent_id_col).value).split('.')[-1] == str(max_sent_num)):
                    edf_row2 += 1
                while(edfws.cell(row=edf_row2,column=rsl_col).value != None):
                    onset = edfws.cell(row=edf_row2,column=sent_onset_col).value
                    if onset != None and str(onset) != '.':
                        onset = float(onset)/1000
                        if onset != 0.0 and onset not in question_onsets:
                            question_onsets.append(onset)
                    offset = edfws.cell(row=edf_row2,column=sent_offset_col).value
                    if offset != None and str(offset) != '.':
                        offset = float(offset)/1000
                        if offset != 0.0 and offset not in question_offsets:
                            question_offsets.append(offset)
                    edf_row2 += 1
                if len(question_onsets) != len(question_offsets) or len(question_onsets) != 10:
                    print "There was a problem determining the onsets and offets of the questions."
                else:
                    question_durations = [question_offsets[i] - question_onsets[i] for i in range(len(question_onsets))]
                    questions = ''
                    for i in range(len(question_onsets)):
                        questions += "{0:.2f}".format(question_onsets[i]) + '\t' "{0:.2f}".format(question_durations[i]) + '\t1.0\n'
                    questions = questions.rstrip('\n')
                    open(''.join([EV_DIR,'Run',str(runnum),'/','questions.run00',str(runnum),'.txt']),'w').write(questions)      
                
            if not os.path.isfile(''.join([EV_DIR,'Run',str(runnum),'/','instructions.run00',str(runnum),'.txt'])):
                instructions = ''
                if not os.path.isfile(''.join([EV_DIR,'Run',str(runnum),'/','first_instructions.run00',str(runnum),'.txt'])):
                    print "Could not make complete instructions file."
                else:
                    instructions += open(''.join([EV_DIR,'Run',str(runnum),'/','first_instructions.run00',str(runnum),'.txt']),'r').read() + '\n' 
                    if not os.path.isfile(''.join([EV_DIR,'Run',str(runnum),'/','text_end_time.run00',str(runnum),'.txt'])):
                        print "Could not make complete instructions file."
                    else:
                        instructions += open(''.join([EV_DIR,'Run',str(runnum),'/','text_end_time.run00',str(runnum),'.txt']),'r').read() + '\t5.0\t1.0\n'
                        if not os.path.isfile(''.join([EV_DIR,'Run',str(runnum),'/','questions.run00',str(runnum),'.txt'])):
                            print "Could not make complete instructions file."
                        else:
                            instructions += open(''.join([EV_DIR,'Run',str(runnum),'/','questions.run00',str(runnum),'.txt']),'r').readlines()[-1].split('\t')[0] + '\t3.0\t1.0'
                            open(''.join([EV_DIR,'Run',str(runnum),'/','instructions.run00',str(runnum),'.txt']),'w').write(instructions)
                            
            previous_fix_id = 0
            current_fix_id = 0
            farthest_fix_id = 0
            previous_sent_id = 0
            current_sent_id = 0
            
            while(edfws.cell(row=edf_row,column=rsl_col).value!=None):
                sent_id = str(edfws.cell(row=edf_row,column = sent_id_col).value)
                if not(sent_id==str(0) or sent_id=='.'):
                    area_id = str(edfws.cell(row=edf_row,column = area_id_col).value)
                    if area_id != str(0) and area_id != '.' and area_id != None and not (sent_id == 't.04.18' and area_id == str(14)):   
                        previous_sent_id = current_sent_id
                        current_sent_id = int(sent_id.split('.')[2])
                        fix_rel_onset = int(edfws.cell(row=edf_row,column = fix_onset_col).value)
                        sent_dur = int(edfws.cell(row=edf_row,column = sent_dur_col).value)
                        if sent_dur == 0:
                            sent_dur = int((float(edfws.cell(row=edf_row,column = sent_offset_col).value) - float(edfws.cell(row=edf_row,column = sent_onset_col).value)))
                        
                        if not(fix_rel_onset>sent_dur):
                
                            true_onset = float(edfws.cell(row=edf_row,column = sent_onset_col).value)
                            true_onset += float(edfws.cell(row=edf_row,column = fix_onset_col).value)
                            true_onset *= (1.0/1000.0)
                            
                            fix_duration = float(edfws.cell(row=edf_row,column = fix_dur_col).value)/1000.0
                            
                            
                            if current_sent_id == previous_sent_id:
                                previous_fix_id = current_fix_id
                                current_fix_id = int(edfws.cell(row=edf_row,column=area_id_col).value)
                                farthest_fix_id = max(previous_fix_id,current_fix_id,farthest_fix_id)
                                if current_fix_id < farthest_fix_id:
                                    regression = 1
                                else:
                                    regression = 0
                            else:
				print "New sentence!"
                                previous_fix_id = 0
                                current_fix_id = int(edfws.cell(row=edf_row,column=area_id_col).value)
                                farthest_fix_id = current_fix_id
                                regression = 0
                                    
                            textdata_row = 1
                            while(text_data.cell(row=textdata_row,column=3).value!=None and 
                                  str(text_data.cell(row=textdata_row,column=3).value)!=sent_id):
                                textdata_row += 1
                            
                            while(text_data.cell(row=textdata_row,column=4).value!=None and 
                                  int(text_data.cell(row=textdata_row,column=4).value)!= int(area_id)):
                                textdata_row += 1
                                
                            w = text_data.cell(row = textdata_row,column=1).value
                            l = text_data.cell(row = textdata_row,column=2).value
                            if l == None or l == '':
                                l = w
            
                            if u'\u201c' in w:
                                w = str(w.replace(u'\u201c','"'))
                            if u'\u201d' in w:
                                w = str(w.replace(u'\u201d','"'))
                            w = charsonly(w).lower()
                            
                            if u'\u201c' in l:
                                l = str(l.replace(u'\u201c','"'))
                            if u'\u201d' in l:
                                l = str(l.replace(u'\u201d','"'))
                            l = charsonly(l).lower()
                            
                            while len(w)<20:
                                w += " "
                            while len(l)<20:
                                l += " "
                            print current_sent_id,current_fix_id,farthest_fix_id,regression
                            #Check for multiple, consecutive fixations on a word to convert to first, second, third, ... pass reading time
                            if current_sent_id == previous_sent_id and current_fix_id == previous_fix_id:
                                if regression == 1:
                                    
                                    key = 'Regressions'
                                    keyww = key + '_with_Words'
                                    evs[key] = evs[key].rstrip('\n').split('\n')
                                    evs[key][-1] = '\t'.join([evs[key][-1].split('\t')[0],"{0:.2f}".format(float(evs[key][-1].split('\t')[1]) + fix_duration),evs[key][-1].split('\t')[2]])
                                    evs[key] = '\n'.join(evs[key]) + '\n'
                                    evs[keyww] = evs[keyww].rstrip('\n').split('\n')
                                    evs[keyww][-1] = '\t'.join([evs[keyww][-1].split('\t')[0],evs[keyww][-1].split('\t')[1],evs[keyww][-1].split('\t')[2],"{0:.2f}".format(float(evs[keyww][-1].split('\t')[2])*-1 + true_onset + fix_duration),evs[keyww][-1].split('\t')[4]])
                                    evs[keyww] = '\n'.join(evs[keyww]) + '\n'
                                    
                                    if int(text_data.cell(row=textdata_row,column=8).value) == int(1):
                                        for key in ['No_Interest','No_Interest_Regressions']:
                                            keyww = key + '_with_Words'
                                            evs[key] = evs[key].rstrip('\n').split('\n')
                                            evs[key][-1] = '\t'.join([evs[key][-1].split('\t')[0],"{0:.2f}".format(float(evs[key][-1].split('\t')[1]) + fix_duration),evs[key][-1].split('\t')[2]])
                                            evs[key] = '\n'.join(evs[key]) + '\n'
                                            evs[keyww] = evs[keyww].rstrip('\n').split('\n')
                                            evs[keyww][-1] = '\t'.join([evs[keyww][-1].split('\t')[0],evs[keyww][-1].split('\t')[1],evs[keyww][-1].split('\t')[2],"{0:.2f}".format(float(evs[keyww][-1].split('\t')[2])*-1 + true_onset + fix_duration),evs[keyww][-1].split('\t')[4]])
                                            evs[keyww] = '\n'.join(evs[keyww]) + '\n'
                                        
                                    if int(text_data.cell(row=textdata_row,column=9).value) == int(1):
                                        for key in ['Target_Word_with_Regressions','Content_or_Target_Word_with_Regressions','Target_Word_Regressions','Content_or_Target_Word_Regressions']:
                                            keyww = key + '_with_Words'
                                            evs[key] = evs[key].rstrip('\n').split('\n')
                                            evs[key][-1] = '\t'.join([evs[key][-1].split('\t')[0],"{0:.2f}".format(float(evs[key][-1].split('\t')[1]) + fix_duration),evs[key][-1].split('\t')[2]])
                                            evs[key] = '\n'.join(evs[key]) + '\n'
                                            evs[keyww] = evs[keyww].rstrip('\n').split('\n')
                                            evs[keyww][-1] = '\t'.join([evs[keyww][-1].split('\t')[0],evs[keyww][-1].split('\t')[1],evs[keyww][-1].split('\t')[2],"{0:.2f}".format(float(evs[keyww][-1].split('\t')[2])*-1 + true_onset + fix_duration),evs[keyww][-1].split('\t')[4]])
                                            evs[keyww] = '\n'.join(evs[keyww]) + '\n'
                                    
                                        for key in evs.keys():
                                            if key not in binaries and 'with_Words' not in key and 'Target' in key and 'with_Regressions' in key:
                                                evs[key]['text'] = evs[key]['text'].rstrip('\n').split('\n')
                                                evs[key]['text'][-1] = '\t'.join([evs[key]['text'][-1].split('\t')[0],"{0:.2f}".format(float(evs[key]['text'][-1].split('\t')[1]) + fix_duration),evs[key]['text'][-1].split('\t')[2]])
                                                evs[key]['text'] = '\n'.join(evs[key]['text']) + '\n'
                                                
                                    if int(text_data.cell(row=textdata_row,column=5).value) == int(1):
                                        for key in ['Content_Word_with_Regressions','Content_or_Target_Word_with_Regressions','Content_Word_Regressions','Content_or_Target_Word_Regressions']:
                                            keyww = key + '_with_Words'
                                            evs[key] = evs[key].rstrip('\n').split('\n')
                                            evs[key][-1] = '\t'.join([evs[key][-1].split('\t')[0],"{0:.2f}".format(float(evs[key][-1].split('\t')[1]) + fix_duration),evs[key][-1].split('\t')[2]])
                                            evs[key] = '\n'.join(evs[key]) + '\n'
                                            evs[keyww] = evs[keyww].rstrip('\n').split('\n')
                                            evs[keyww][-1] = '\t'.join([evs[keyww][-1].split('\t')[0],evs[keyww][-1].split('\t')[1],evs[keyww][-1].split('\t')[2],"{0:.2f}".format(float(evs[keyww][-1].split('\t')[2])*-1 + true_onset + fix_duration),evs[keyww][-1].split('\t')[4]])
                                            evs[keyww] = '\n'.join(evs[keyww]) + '\n'
                                
                                        for key in evs.keys():
                                            if key not in binaries and 'with_Words' not in key and 'Content' in key and 'with_Regressions' in key:
                                                evs[key]['text'] = evs[key]['text'].rstrip('\n').split('\n')
                                                evs[key]['text'][-1] = '\t'.join([evs[key]['text'][-1].split('\t')[0],"{0:.2f}".format(float(evs[key]['text'][-1].split('\t')[1]) + fix_duration),evs[key]['text'][-1].split('\t')[2]])
                                                evs[key]['text'] = '\n'.join(evs[key]['text']) + '\n'
                                                
                                else:
                                    
                                    if int(text_data.cell(row=textdata_row,column=8).value) == int(1):
                                        for key in ['No_Interest','No_Interest_without_Regressions']:
                                            keyww = key + '_with_Words'
                                            evs[key] = evs[key].rstrip('\n').split('\n')
                                            evs[key][-1] = '\t'.join([evs[key][-1].split('\t')[0],"{0:.2f}".format(float(evs[key][-1].split('\t')[1]) + fix_duration),evs[key][-1].split('\t')[2]])
                                            evs[key] = '\n'.join(evs[key]) + '\n'
                                            evs[keyww] = evs[keyww].rstrip('\n').split('\n')
                                            evs[keyww][-1] = '\t'.join([evs[keyww][-1].split('\t')[0],evs[keyww][-1].split('\t')[1],evs[keyww][-1].split('\t')[2],"{0:.2f}".format(float(evs[keyww][-1].split('\t')[2])*-1 + true_onset + fix_duration),evs[keyww][-1].split('\t')[4]])
                                            evs[keyww] = '\n'.join(evs[keyww]) + '\n'
                                        
                                    if int(text_data.cell(row=textdata_row,column=9).value) == int(1):
                                        for key in ['Target_Word_with_Regressions','Content_or_Target_Word_with_Regressions','Target_Word','Content_or_Target_Word']:
                                            keyww = key + '_with_Words'
                                            evs[key] = evs[key].rstrip('\n').split('\n')
                                            evs[key][-1] = '\t'.join([evs[key][-1].split('\t')[0],"{0:.2f}".format(float(evs[key][-1].split('\t')[1]) + fix_duration),evs[key][-1].split('\t')[2]])
                                            evs[key] = '\n'.join(evs[key]) + '\n'
                                            evs[keyww] = evs[keyww].rstrip('\n').split('\n')
                                            evs[keyww][-1] = '\t'.join([evs[keyww][-1].split('\t')[0],evs[keyww][-1].split('\t')[1],evs[keyww][-1].split('\t')[2],"{0:.2f}".format(float(evs[keyww][-1].split('\t')[2])*-1 + true_onset + fix_duration),evs[keyww][-1].split('\t')[4]])
                                            evs[keyww] = '\n'.join(evs[keyww]) + '\n'
                                        
                                        for key in evs.keys():
                                            if key not in binaries and 'with_Words' not in key and 'Target' in key:
                                                evs[key]['text'] = evs[key]['text'].rstrip('\n').split('\n')
                                                evs[key]['text'][-1] = '\t'.join([evs[key]['text'][-1].split('\t')[0],"{0:.2f}".format(float(evs[key]['text'][-1].split('\t')[1]) + fix_duration),evs[key]['text'][-1].split('\t')[2]])
                                                evs[key]['text'] = '\n'.join(evs[key]['text']) + '\n'
                                        
                                    if int(text_data.cell(row=textdata_row,column=5).value) == int(1):
                                        for key in ['Content_Word_with_Regressions','Content_or_Target_Word_with_Regressions','Content_Word','Content_or_Target_Word']:
                                            keyww = key + '_with_Words'
                                            evs[key] = evs[key].rstrip('\n').split('\n')
                                            evs[key][-1] = '\t'.join([evs[key][-1].split('\t')[0],"{0:.2f}".format(float(evs[key][-1].split('\t')[1]) + fix_duration),evs[key][-1].split('\t')[2]])
                                            evs[key] = '\n'.join(evs[key]) + '\n'
                                            evs[keyww] = evs[keyww].rstrip('\n').split('\n')
                                            evs[keyww][-1] = '\t'.join([evs[keyww][-1].split('\t')[0],evs[keyww][-1].split('\t')[1],evs[keyww][-1].split('\t')[2],"{0:.2f}".format(float(evs[keyww][-1].split('\t')[2])*-1 + true_onset + fix_duration),evs[keyww][-1].split('\t')[4]])
                                            evs[keyww] = '\n'.join(evs[keyww]) + '\n'
            
                                        for key in evs.keys():
                                            if key not in binaries and 'with_Words' not in key and 'Content' in key:
                                                evs[key]['text'] = evs[key]['text'].rstrip('\n').split('\n')
                                                evs[key]['text'][-1] = '\t'.join([evs[key]['text'][-1].split('\t')[0],"{0:.2f}".format(float(evs[key]['text'][-1].split('\t')[1]) + fix_duration),evs[key]['text'][-1].split('\t')[2]])
                                                evs[key]['text'] = '\n'.join(evs[key]['text']) + '\n'
                                                
                            #Add a new fixation event
                            else:
                                if regression == 1:
                                    
                                    val = "{0:.2f}".format(float(1)) + '\n'
                                    evs['Regressions'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                    evs['Regressions_with_Words'] += '\t'.join([w,l,"{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                    
                                    if int(text_data.cell(row=textdata_row,column=8).value) == int(1):
                                        val = "{0:.2f}".format(float(1)) + '\n'
                                        evs['No_Interest'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['No_Interest_with_Words'] += '\t'.join([w,l,"{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['No_Interest_Regressions'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['No_Interest_Regressions_with_Words'] += '\t'.join([w,l,"{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        
                                    if int(text_data.cell(row=textdata_row,column=9).value) == int(1):
                                        val = "{0:.2f}".format(float(1)) + '\n'
                                        evs['Target_Word_with_Regressions'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Target_Word_with_Regressions_with_Words'] += '\t'.join([w,l,"{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_or_Target_Word_with_Regressions'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_or_Target_Word_with_Regressions_with_Words'] += '\t'.join([w,l,"{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Target_Word_Regressions'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Target_Word_Regressions_with_Words'] += '\t'.join([w,l,"{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_or_Target_Word_Regressions'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_or_Target_Word_Regressions_with_Words'] += '\t'.join([w,l,"{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                    
                                        for key in evs.keys():
                                            if key not in binaries and 'with_Words' not in key and 'Target' in key and 'with_Regressions' in key:
                                                val = "{0:.2f}".format(float(text_data.cell(row=textdata_row,column=evs[key]['column']).value)) + '\n'
                                                evs[key]['text'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                                
                                    if int(text_data.cell(row=textdata_row,column=5).value) == int(1):
                                        val = "{0:.2f}".format(float(1)) + '\n'
                                        evs['Content_Word_with_Regressions'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_Word_with_Regressions_with_Words'] += '\t'.join([w,l,"{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_or_Target_Word_with_Regressions'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_or_Target_Word_with_Regressions_with_Words'] += '\t'.join([w,l,"{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_Word_Regressions'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_Word_Regressions_with_Words'] += '\t'.join([w,l,"{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_or_Target_Word_Regressions'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_or_Target_Word_Regressions_with_Words'] += '\t'.join([w,l,"{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                
                                        for key in evs.keys():
                                            if key not in binaries and 'with_Words' not in key and 'Content' in key and 'with_Regressions' in key:
                                                val = "{0:.2f}".format(float(text_data.cell(row=textdata_row,column=evs[key]['column']).value)) + '\n'
                                                evs[key]['text'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                                
                                else:
                                    
                                    if int(text_data.cell(row=textdata_row,column=8).value) == int(1):
                                        val = "{0:.2f}".format(float(1)) + '\n'
                                        evs['No_Interest'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['No_Interest_with_Words'] += '\t'.join([w,l,"{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['No_Interest_without_Regressions'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['No_Interest_without_Regressions_with_Words'] += '\t'.join([w,l,"{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        
                                    if int(text_data.cell(row=textdata_row,column=9).value) == int(1):
                                        val = "{0:.2f}".format(float(1)) + '\n'
                                        evs['Target_Word_with_Regressions'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Target_Word_with_Regressions_with_Words'] += '\t'.join([w,l,"{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_or_Target_Word_with_Regressions'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_or_Target_Word_with_Regressions_with_Words'] += '\t'.join([w,l,"{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Target_Word'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Target_Word_with_Words'] += '\t'.join([w,l,"{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_or_Target_Word'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_or_Target_Word_with_Words'] += '\t'.join([w,l,"{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        
                                        for key in evs.keys():
                                            if key not in binaries and 'with_Words' not in key and 'Target' in key:
                                                val = "{0:.2f}".format(float(text_data.cell(row=textdata_row,column=evs[key]['column']).value)) + '\n'
                                                evs[key]['text'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                    
                                        
                                    if int(text_data.cell(row=textdata_row,column=5).value) == int(1):
                                        val = "{0:.2f}".format(float(1)) + '\n'
                                        evs['Content_Word_with_Regressions'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_Word_with_Regressions_with_Words'] += '\t'.join([w,l,"{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_or_Target_Word_with_Regressions'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_or_Target_Word_with_Regressions_with_Words'] += '\t'.join([w,l,"{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_Word'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_Word_with_Words'] += '\t'.join([w,l,"{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_or_Target_Word'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        evs['Content_or_Target_Word_with_Words'] += '\t'.join([w,l,"{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                                        
                                        for key in evs.keys():
                                            if key not in binaries and 'with_Words' not in key and 'Content' in key:
                                                val = "{0:.2f}".format(float(text_data.cell(row=textdata_row,column=evs[key]['column']).value)) + '\n'
                                                evs[key]['text'] += '\t'.join(["{0:.2f}".format(true_onset),"{0:.2f}".format(fix_duration),val])
                edf_row += 1
            
            for key in evs.keys():
                if key in binaries or 'with_Words' in key:                
                    evs[key] = evs[key].rstrip('\n')
                    evs['_'.join([key,'_with_Dummy'])] = evs[key]
                    evs[key] = evs[key].split('\n')
                else:
                    evs[key]['text'] = evs[key]['text'].rstrip('\n')
                    evs[key]['text'] = evs[key]['text'].split('\n') 
                               
            #Remove lines with missing values and put in dummy variable NOT INCLUDING regressions
            for suffix in ['Target','Content','Content_or_Target']:
                evs['_'.join(['Dummy',suffix])] = ''
                evs['_'.join(['Dummy',suffix,'with_Words'])] = ''
                remove_lines = []
                for i in range(len(evs['_'.join(['AoA',suffix])]['text'])):
                    zero_vals = 0
                    for key in evs.keys():
                        if key not in binaries and 'with_Words' not in key and 'Dummy' not in key and 'Imageability' not in key and 'with_Regressions' not in key:
                            if suffix == 'Content_or_Target' and 'Content_or_Target' in key:
                                line = evs[key]['text'][i].split('\t')
                                if float(line[-1].rstrip('\n')) == float(0):
                                    zero_vals += 1
                            elif suffix == 'Content' and key[-7:] == 'Content':
                                line = evs[key]['text'][i].split('\t')
                                if float(line[-1].rstrip('\n')) == float(0):
                                    zero_vals += 1
                            elif suffix == 'Target' and key[-6:] == 'Target' and 'Content_or_Target' not in key:
                                line = evs[key]['text'][i].split('\t')
                                if float(line[-1].rstrip('\n')) == float(0):
                                    zero_vals += 1
                    if zero_vals > 0:
                        remove_lines.append(i)
                remove_lines = sorted(remove_lines,reverse=True)
                for line in remove_lines:
                    onset = evs['_'.join(['AoA',suffix])]['text'][line].split('\t')[0]
                    duration = evs['_'.join(['AoA',suffix])]['text'][line].split('\t')[1]
                    val = "{0:.2f}".format(float(1)) + '\n'
                    evs['_'.join(['Dummy',suffix])] += '\t'.join([onset,duration,val])
                    evs['_'.join(['Dummy',suffix,'with_Words'])] += evs['_'.join([suffix,'Word','with_Words'])][line] + '\n'
                    for key in evs.keys():
                        if key not in binaries and 'with_Words' not in key and 'Dummy' not in key and 'with_Regressions' not in key:
                            if suffix == 'Content_or_Target' and 'Content_or_Target' in key:
                                del evs[key]['text'][line] 
                            elif suffix == 'Content' and key[-7:] == 'Content':
                                del evs[key]['text'][line]
                            elif suffix == 'Target' and key[-6:] == 'Target' and 'Content_or_Target' not in key:
                                del evs[key]['text'][line]
                    del evs['_'.join([suffix,'Word'])][line]
                    del evs['_'.join([suffix,'Word','with_Words'])][line]
            
            #Remove lines with missing values and put in dummy variable INCLUDING regressions
            for suffix in ['Target_with_Regressions','Content_with_Regressions','Content_or_Target_with_Regressions']:
                evs['_'.join(['Dummy',suffix.replace('_with','_Word_with')])] = ''
                evs['_'.join(['Dummy',suffix.replace('_with','_Word_with'),'with_Words'])] = ''
                remove_lines = []
                for i in range(len(evs['_'.join(['AoA',suffix])]['text'])):
                    zero_vals = 0
                    for key in evs.keys():
                        if key not in binaries and 'with_Words' not in key and 'Dummy' not in key and 'Imageability' not in key and 'with_Regressions' in key:
                            if suffix == 'Content_or_Target_with_Regressions' and 'Content_or_Target_with_Regressions' in key:
                                line = evs[key]['text'][i].split('\t')
                                if float(line[-1].rstrip('\n')) == float(0):
                                    zero_vals += 1
                            elif suffix == 'Content_with_Regressions' and key[-24:] == 'Content_with_Regressions':
                                line = evs[key]['text'][i].split('\t')
                                if float(line[-1].rstrip('\n')) == float(0):
                                    zero_vals += 1
                            elif suffix == 'Target_with_Regressions' and key[-23:] == 'Target_with_Regressions' and 'Content_or_Target_with_Regressions' not in key:
                                line = evs[key]['text'][i].split('\t')
                                if float(line[-1].rstrip('\n')) == float(0):
                                    zero_vals += 1
                    if zero_vals > 0:
                        remove_lines.append(i)
                remove_lines = sorted(remove_lines,reverse=True)
                for line in remove_lines:
                    onset = evs['_'.join(['AoA',suffix])]['text'][line].split('\t')[0]
                    duration = evs['_'.join(['AoA',suffix])]['text'][line].split('\t')[1]
                    val = "{0:.2f}".format(float(1)) + '\n'
                    evs['_'.join(['Dummy',suffix.replace('_with','_Word_with')])] += '\t'.join([onset,duration,val])
                    evs['_'.join(['Dummy',suffix.replace('_with','_Word_with'),'with_Words'])] += evs['_'.join([suffix.replace('_with','_Word_with'),'with_Words'])][line] + '\n'
                    for key in evs.keys():
                        if key not in binaries and 'with_Words' not in key and 'Dummy' not in key and 'with_Regressions' in key:
                            if suffix == 'Content_or_Target_with_Regressions' and 'Content_or_Target_with_Regressions' in key:
                                del evs[key]['text'][line] 
                            elif suffix == 'Content_with_Regressions' and key[-24:] == 'Content_with_Regressions':
                                del evs[key]['text'][line]
                            elif suffix == 'Target_with_Regressions' and key[-23:] == 'Target_with_Regressions' and 'Content_or_Target_with_Regressions' not in key:
                                del evs[key]['text'][line]
                    del evs[suffix.replace('_with','_Word_with')][line]
                    del evs['_'.join([suffix.replace('_with','_Word_with'),'with_Words'])][line]     
                 
            for key in evs.keys():
                if 'Dummy' in key:
                    open(''.join([EV_DIR,'Run',str(runnum),'/',key,'.run00',str(runnum),'.txt']),'w').write(evs[key].rstrip('\n'))
                elif key in binaries or 'with_Words' in key:
                    if(l2t(evs[key])!='\n' and l2t(evs[key])!=''):
                        open(''.join([EV_DIR,'Run',str(runnum),'/',key,'.run00',str(runnum),'.txt']),'w').write(l2t(evs[key]))
                        #open(''.join([EV_DIR,'Run',str(runnum),'/',key,'_with_Words.run00',str(runnum),'.txt']),'w').write(l2t(evs['_'.join([key,'with_Words'])]))
                    else:
                        open(''.join([EV_DIR,'Run',str(runnum),'/',key,'.run00',str(runnum),'.txt']),'w').write('0.00\t0.00\t0.00')   
                        #open(''.join([EV_DIR,'Run',str(runnum),'/',key,'_with_Words.run00',str(runnum),'.txt']),'w').write('None\t0.00\t0.00\t0.00')   
                else:
                    open(''.join([EV_DIR,'Run',str(runnum),'/',key,'.run00',str(runnum),'.txt']),'w').write(l2t(normalize(evs[key]['text'])))
                    open(''.join([EV_DIR,'Run',str(runnum),'/',key,'_Squared.run00',str(runnum),'.txt']),'w').write(l2t(quadratic(normalize(evs[key]['text']))))

if __name__=='__main__':    
    main()    
