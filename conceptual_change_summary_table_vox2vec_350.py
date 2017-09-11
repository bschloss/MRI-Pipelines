# -*- coding: utf-8 -*-
"""
Created on Wed Feb 15 15:40:57 2017

@author: bschloss
"""
import openpyxl as xl
import scipy.io as sio
import numpy as np

models = ['context','order','contextorder','sg','bow','sgbow','contextsgbow','contextordersgbow']
#pars = ['201','002','003','004','105','006','107','008','009','110',
#        '011','012','013','214','015','016','017','018','019','020',
#        '021','122','023','024','025','026','027','028','029','030',
#        '031','132','033','034','035','036','037','038','039','040',
#        '041','042','043','044','045','046','047','048','049','050']
pars = ['2101','2102','2003','2104','2005','2006','2007','2008','2009','2010',
           '2011','2012','2013','2014','2015','2016','2017','2218','2019','2020',
           '2021','2022','2123','2024','2025','2026','2027','2128']
wb = xl.Workbook()
ws = wb.active
for row in range(2,len(pars)+2):
     ws.cell(row=row,column=1).value = pars[row-2][1:]
for col in range(2,len(models)+2):
     ws.cell(row=1,column=col).value = models[col-2] 
     
results = sio.loadmat('/gpfs/group/pul8/default/read/Group_Analyses/Conceptual_Change_Bilingual/results_svd_sv2_350_all_participants_vox2vec.mat')

for i in range(len(models)):
    model = models[i]
    for j in range(len(pars)):
        val = "{0:.2f}".format(results[model][j,0])
        if results[model][j,1] < .05:
            val += '*'
        if results[model][j,2] < .05:
            val += '*'
        ws.cell(row=j+2,column=i+2).value = val
row = len(pars) + 3
ws.cell(row=row,column=1).value = 'Avg*/NumSig'
for i in range(len(models)):
    model = models[i]
    logi = results[model][:,1]<.05
    avg = np.mean(results[model][logi,0])
    val = "{0:.2f}".format(avg) + '/' + "{0:.2f}".format(sum(logi)/float(len(logi)))
    ws.cell(row=row,column=i+2).value = val
row += 1   
ws.cell(row=row,column=1).value = 'Avg**/NumSig'
for i in range(len(models)):
    model = models[i]
    logi = results[model][:,2]<.05
    avg = np.mean(results[model][logi,0])
    val = "{0:.2f}".format(avg) + '/' + "{0:.2f}".format(sum(logi)/float(len(logi)))
    ws.cell(row=row,column=i+2).value = val
    
wb.save('/gpfs/group/pul8/default/read/Group_Analyses/Conceptual_Change_Bilingual/summary_svd_sv2_350_vox2vec.xlsx')
